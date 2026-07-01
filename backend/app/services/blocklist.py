"""Generate exportable blocklist (xlsx) from classified attackers.

Produces a 3-sheet workbook:
  1. "Atacantes" — full per-IP detail.
  2. "Bloqueos /24" — aggregated by subnet when >= 2 IPs share /24.
  3. "Pegar en USG" — minimal list of CIDRs + reason, ordered by priority.
"""
import ipaddress
import time
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from . import geoip as geo_svc
from . import security as sec_svc


CATEGORY_LABELS = {
    "attack": "ATAQUE",
    "scan": "ESCANEO",
    "service": "SERVICIO",
    "noise": "RUIDO",
}

CATEGORY_FILL = {
    "attack": PatternFill("solid", fgColor="FFEBEB"),  # rojo claro
    "scan": PatternFill("solid", fgColor="FFF4E3"),  # ámbar claro
    "service": PatternFill("solid", fgColor="F2E8FA"),  # púrpura claro
    "noise": PatternFill("solid", fgColor="F1F4F6"),  # gris claro
}

CATEGORY_FONT_COLOR = {
    "attack": "B40000",
    "scan": "A86200",
    "service": "5C2D8F",
    "noise": "6B7B86",
}

HEADER_FILL = PatternFill("solid", fgColor="006876")  # primary
HEADER_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")


def _classify_filter(category_filter: str):
    """Returns a predicate function for filtering rows by category."""
    if category_filter == "real":
        return lambda r: r.get("category") in ("attack", "scan", "service")
    if category_filter == "all":
        return lambda r: True
    return lambda r: r.get("category") == category_filter


def _subnet_24(ip: str) -> str:
    try:
        return str(ipaddress.ip_network(f"{ip}/24", strict=False).network_address) + "/24"
    except ValueError:
        return f"{ip}/32"


async def build_blocklist_xlsx(
    range_s: int,
    category_filter: str = "real",
    min_attempts: int = 0,
) -> bytes:
    rows = await sec_svc.attack_summary(range_s, limit=500)
    geo = await geo_svc.lookup_many([r["src_ip"] for r in rows])
    for r in rows:
        g = geo.get(r["src_ip"]) or {}
        r["country"] = g.get("country")
        r["country_code"] = g.get("country_code")
        r["city"] = g.get("city")
        r["isp"] = g.get("isp")
        r["asn"] = g.get("asn")

    pred = _classify_filter(category_filter)
    filtered = [r for r in rows if pred(r) and (r.get("attempts") or 0) >= min_attempts]
    filtered.sort(key=lambda r: r.get("score", 0), reverse=True)

    wb = Workbook()

    # ===== Sheet 1: Atacantes =====
    ws = wb.active
    ws.title = "Atacantes"
    cols = [
        ("Categoría", 12),
        ("IP", 18),
        ("Bloque /24", 18),
        ("País", 8),
        ("Ciudad", 16),
        ("ISP / Organización", 32),
        ("ASN", 14),
        ("Intentos", 10),
        ("Puertos distintos", 12),
        ("Top puertos", 32),
        ("Primer visto", 18),
        ("Última vez", 18),
        ("Score", 8),
        ("Motivo clasificación", 50),
    ]
    for i, (header, width) in enumerate(cols, start=1):
        c = ws.cell(row=1, column=i, value=header)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.column_dimensions[get_column_letter(i)].width = width

    for ridx, r in enumerate(filtered, start=2):
        cat = r.get("category", "noise")
        fill = CATEGORY_FILL[cat]
        cat_font = Font(bold=True, color=CATEGORY_FONT_COLOR[cat], name="Calibri", size=10)

        top_ports = ", ".join(f"{p['dst_port']}({p['n']})" for p in r.get("top_ports", [])[:5])
        first_seen = r.get("first_seen")
        last_seen = r.get("last_seen")
        first_str = time.strftime("%Y-%m-%d %H:%M", time.localtime(first_seen)) if first_seen else ""
        last_str = time.strftime("%Y-%m-%d %H:%M", time.localtime(last_seen)) if last_seen else ""

        values = [
            CATEGORY_LABELS[cat],
            r["src_ip"],
            _subnet_24(r["src_ip"]),
            r.get("country_code") or "",
            r.get("city") or "",
            r.get("isp") or "",
            r.get("asn") or "",
            r.get("attempts") or 0,
            r.get("distinct_ports") or 0,
            top_ports,
            first_str,
            last_str,
            r.get("score") or 0,
            r.get("category_reason") or "",
        ]
        for cidx, v in enumerate(values, start=1):
            c = ws.cell(row=ridx, column=cidx, value=v)
            c.fill = fill
            if cidx == 1:
                c.font = cat_font
                c.alignment = Alignment(horizontal="center")

    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}{max(2, len(filtered) + 1)}"

    # ===== Sheet 2: Bloqueos /24 (only subnets with >=2 attackers) =====
    ws2 = wb.create_sheet("Bloqueos por subnet")
    by_24: dict[str, list[dict]] = {}
    for r in filtered:
        s = _subnet_24(r["src_ip"])
        by_24.setdefault(s, []).append(r)
    subnet_rows = [
        {
            "subnet": s,
            "count": len(items),
            "total_attempts": sum(int(i.get("attempts") or 0) for i in items),
            "categories": ", ".join(sorted({CATEGORY_LABELS[i.get("category", "noise")] for i in items})),
            "country": items[0].get("country_code") or "",
            "isp": items[0].get("isp") or "",
            "ips": ", ".join(i["src_ip"] for i in items),
        }
        for s, items in by_24.items()
        if len(items) >= 2
    ]
    subnet_rows.sort(key=lambda x: (-x["count"], -x["total_attempts"]))

    sub_cols = [
        ("Subnet /24", 18),
        ("# IPs", 8),
        ("Total intentos", 14),
        ("Categorías", 24),
        ("País", 8),
        ("ISP", 32),
        ("IPs incluidas", 60),
    ]
    for i, (h, w) in enumerate(sub_cols, start=1):
        c = ws2.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        ws2.column_dimensions[get_column_letter(i)].width = w

    for ridx, sr in enumerate(subnet_rows, start=2):
        for cidx, v in enumerate([
            sr["subnet"], sr["count"], sr["total_attempts"],
            sr["categories"], sr["country"], sr["isp"], sr["ips"],
        ], start=1):
            ws2.cell(row=ridx, column=cidx, value=v)
    ws2.freeze_panes = "A2"

    # ===== Sheet 3: Pegar en USG (ready-to-block) =====
    ws3 = wb.create_sheet("Pegar en USG")
    pegar_cols = [("CIDR a bloquear", 22), ("Motivo", 40), ("Intentos", 10), ("Categoría", 12)]
    for i, (h, w) in enumerate(pegar_cols, start=1):
        c = ws3.cell(row=1, column=i, value=h)
        c.font = HEADER_FONT
        c.fill = HEADER_FILL
        ws3.column_dimensions[get_column_letter(i)].width = w

    # Subnets with multiple attackers go as /24, single attackers as /32
    paste_rows: list[tuple[str, str, int, str]] = []
    seen_subnets: set[str] = set()
    for sr in subnet_rows:
        paste_rows.append((
            sr["subnet"],
            f"{sr['count']} atacantes desde {sr['country']} {sr['isp']}"[:80],
            sr["total_attempts"],
            sr["categories"],
        ))
        seen_subnets.add(sr["subnet"])
    for r in filtered:
        if r.get("category") == "noise":
            continue
        s24 = _subnet_24(r["src_ip"])
        if s24 in seen_subnets:
            continue
        paste_rows.append((
            f"{r['src_ip']}/32",
            f"{CATEGORY_LABELS[r['category']]}: {r.get('isp') or ''} ({r.get('country_code') or '?'})"[:80],
            int(r.get("attempts") or 0),
            CATEGORY_LABELS[r["category"]],
        ))

    for ridx, (cidr, motivo, n, cat) in enumerate(paste_rows, start=2):
        ws3.cell(row=ridx, column=1, value=cidr)
        ws3.cell(row=ridx, column=2, value=motivo)
        ws3.cell(row=ridx, column=3, value=n)
        ws3.cell(row=ridx, column=4, value=cat)
    ws3.freeze_panes = "A2"

    # ===== Sheet 4: Metadatos =====
    ws4 = wb.create_sheet("Info")
    info = [
        ("Generado", time.strftime("%Y-%m-%d %H:%M:%S")),
        ("Ventana", f"últimos {range_s // 3600}h" if range_s >= 3600 else f"últimos {range_s // 60}min"),
        ("Filtro categoría", category_filter),
        ("Min. intentos", min_attempts),
        ("Atacantes totales", len(filtered)),
        ("Subnets agrupados /24", len(subnet_rows)),
        ("CIDRs sugeridos", len(paste_rows)),
        ("", ""),
        ("Aplicación en USG (manual)", ""),
        ("1.", "Configuration → Object → Address → Add"),
        ("2.", "Para cada CIDR de la hoja 'Pegar en USG':"),
        ("3.", "  - Address Type: SUBNET (si /24) o HOST (si /32)"),
        ("4.", "  - Pegar la IP o CIDR en el campo correspondiente"),
        ("5.", "Configuration → Security Policy → Add una nueva regla DENY"),
        ("6.", "  - From: WAN  To: ZyWALL/LAN"),
        ("7.", "  - Source: el Address Object recién creado"),
        ("8.", "  - Action: deny + log"),
        ("9.", "  - Mover regla ARRIBA de las reglas Allow"),
        ("10.", "Apply."),
    ]
    for ridx, (k, v) in enumerate(info, start=1):
        ws4.cell(row=ridx, column=1, value=k).font = Font(bold=True, name="Calibri")
        ws4.cell(row=ridx, column=2, value=v)
    ws4.column_dimensions["A"].width = 30
    ws4.column_dimensions["B"].width = 70

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
